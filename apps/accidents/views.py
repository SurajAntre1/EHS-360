from multiprocessing import context
from urllib.parse import urlencode
from django.contrib.auth.mixins import LoginRequiredMixin,UserPassesTestMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView,DeleteView
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Q, Count
from django.http import JsonResponse
from apps.organizations.models import *
from .models import *
from .forms import *
from .utils import generate_incident_pdf
from django.http import HttpResponse
from django.db.models.functions import TruncMonth
from django.views.generic import UpdateView, TemplateView
from django.contrib import messages
from django.utils import timezone
from apps.accidents.models import IncidentType
from apps.notifications import *
import datetime
from django.db.models import Q
import json
import openpyxl
from django.shortcuts import render
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from django.conf import settings  
from django.conf.urls.static import static  
from apps.common.image_utils import compress_image

from .forms import IncidentAttachmentForm # <-- Import the new form




class IncidentTypeListView(LoginRequiredMixin, ListView):
    """List all incident types with search functionality"""
    model = IncidentType
    template_name = 'accidents/incident_type_list.html'
    context_object_name = 'incident_types'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = IncidentType.objects.annotate(
            incident_count=Count('incidents')
        ).order_by('name')
        
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class IncidentTypeCreateView(LoginRequiredMixin, CreateView):
    """Create a new incident type"""
    model = IncidentType
    form_class = IncidentTypeForm
    template_name = 'accidents/incident_type_form.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def form_valid(self, form):
        incident_type = form.save(commit=False)
        incident_type.created_by = self.request.user
        incident_type.save()
        messages.success(
            self.request, 
            f'Incident Type "{incident_type.name}" created successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Create'
        return context


class IncidentTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing incident type"""
    model = IncidentType
    form_class = IncidentTypeForm
    template_name = 'accidents/incident_type_form.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def form_valid(self, form):
        messages.success(
            self.request,
            f'Incident Type "{self.object.name}" updated successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Update'
        context['incident_type'] = self.object
        return context


class IncidentTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete an incident type"""
    model = IncidentType
    template_name = 'accidents/incident_type_confirm_delete.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident_count'] = self.object.incidents.count()
        return context
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        incident_count = self.object.incidents.count()
        
        if incident_count > 0:
            messages.error(
                request,
                f'Cannot delete "{self.object.name}". It is being used by {incident_count} incident(s).'
            )
            return redirect('accidents:incident_type_list')
        
        incident_type_name = self.object.name
        success_url = self.get_success_url()
        self.object.delete()
        
        messages.success(
            request,
            f'Incident Type "{incident_type_name}" deleted successfully!'
        )
        return redirect(success_url)
class IncidentDashboardView(LoginRequiredMixin, TemplateView):
    """Incident Management Dashboard"""
    template_name = 'accidents/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get incidents based on user role
        if self.request.user.is_superuser or self.request.user.role.name == 'ADMIN':
            incidents = Incident.objects.all()
        elif self.request.user.plant:
            incidents = Incident.objects.filter(plant=self.request.user.plant)
        else:
            incidents = Incident.objects.filter(reported_by=self.request.user)
        
        # Statistics
        context['total_incidents'] = incidents.count()
        context['open_incidents'] = incidents.exclude(status='CLOSED').count()
        context['this_month_incidents'] = incidents.filter(
            incident_date__month=datetime.date.today().month,
            incident_date__year=datetime.date.today().year
        ).count()
        context['investigation_pending'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True
        ).count()
        
        # ✅ UPDATED: By Type using ForeignKey relationship
        # Get incident type IDs for common types
        try:
            lti_type = IncidentType.objects.get(code='LTI')
            context['lti_count'] = incidents.filter(incident_type__code='LTI').count()
        except IncidentType.DoesNotExist:
            context['lti_count'] = 0
        
        try:
            mtc_type = IncidentType.objects.get(code='MTC')
            context['mtc_count'] = incidents.filter(incident_type__code='MTC').count()
        except IncidentType.DoesNotExist:
            context['mtc_count'] = 0
        
        try:
            fa_type = IncidentType.objects.get(code='FA')
            context['fa_count'] = incidents.filter(incident_type__code='FA').count()
        except IncidentType.DoesNotExist:
            context['fa_count'] = 0
        
        try:
            hlfi_type = IncidentType.objects.get(code='HLFI')
            context['hlfi_count'] = incidents.filter(incident_type__code='HLFI').count()
        except IncidentType.DoesNotExist:
            context['hlfi_count'] = 0
        
        # Recent incidents
        context['recent_incidents'] = incidents.order_by('-reported_date')[:10]
        
        # Overdue investigations
        context['overdue_investigations'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True,
            investigation_deadline__lt=datetime.date.today()
        )
        
        return context


class IncidentListView(LoginRequiredMixin, ListView):
    """List all incidents"""
    model = Incident
    template_name = 'accidents/incident_list.html'
    context_object_name = 'incidents'
    paginate_by = 20
    
    def get_queryset(self):
        # Get the current logged-in user
        user = self.request.user
        
        # Start with the base queryset, fetching related objects to optimize queries
        queryset = Incident.objects.select_related('plant', 'location', 'reported_by','incident_type').order_by('-incident_date', '-incident_time')
        
        # --- ROLE-BASED DATA FILTERING ---
        # Check if the user is a superuser or has an ADMIN role
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            # No filtering needed; they can see all records
            pass
        # Check if the user has an EMPLOYEE role
        elif hasattr(user, 'role') and user.role and user.role.name == 'EMPLOYEE':
            # Filter the queryset to show only records reported by the current user
            queryset = queryset.filter(reported_by=user)
        # Check if the user is associated with a specific plant (for roles like PLANT HEAD, etc.)
        elif user.plant:
            queryset = queryset.filter(plant=user.plant)
        else:
            # As a fallback, if no specific role logic applies, show only self-reported records
            queryset = queryset.filter(reported_by=user)
        
        # --- SEARCH AND FILTER LOGIC ---
        # This part remains the same and applies on top of the role-filtered queryset
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(report_number__icontains=search) |
                Q(affected_person_name__icontains=search)
            )
        
        incident_type = self.request.GET.get('incident_type')
        if incident_type:
            queryset = queryset.filter(incident_type_id=incident_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        plant = self.request.GET.get('plant')
        if plant:
            queryset = queryset.filter(plant_id=plant)
        
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(incident_date__gte=date_from)
            
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(incident_date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.organizations.models import Plant
        context['plants'] = Plant.objects.filter(is_active=True)
        context['incident_types'] = IncidentType.objects.filter(is_active=True).order_by('name')
        context['status_choices'] = Incident.STATUS_CHOICES
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_incident_type'] = self.request.GET.get('incident_type', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_plant'] = self.request.GET.get('plant', '')
        return context


# Updated IncidentCreateView to handle unsafe acts and conditions
class IncidentCreateView(LoginRequiredMixin, CreateView):
    model = Incident
    form_class = IncidentReportForm
    template_name = 'accidents/incident_create.html'
    success_url = reverse_lazy('accidents:incident_list')

    def get_form_kwargs(self):
        """
        Passes the current request's user to the form's __init__ method.
        This is CRUCIAL for the form logic to work.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        """
        Adds the user's location assignments to the template context.
        This allows the template to conditionally render fields as readonly or dropdowns.
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Pass QuerySets of assigned locations to the template
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)
        
        # Pass other assignments for the template logic to use
        # The template can now check the count of each of these
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(is_active=True, location=location)
                else:
                    context['user_assigned_sublocations'] = user.assigned_sublocations.none() # Or all if needed
            else:
                context['user_assigned_locations'] = user.assigned_locations.none()
                context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        else:
            context['user_assigned_zones'] = user.assigned_zones.none()
            context['user_assigned_locations'] = user.assigned_locations.none()
            context['user_assigned_sublocations'] = user.assigned_sublocations.none()

        context['active_incident_types'] = IncidentType.objects.filter(is_active=True)
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        return context
    
    def form_valid(self, form):
        """
        Process the valid form, set the reporter, and handle location data.
        """
        incident = form.save(commit=False)
        incident.reported_by = self.request.user
        
        user = self.request.user

        # Manually set location fields if they are single-assigned and might not be in the form post data
        # (e.g., if we use readonly fields instead of disabled dropdowns).
        if user.assigned_plants.count() == 1 and not form.cleaned_data.get('plant'):
            incident.plant = user.assigned_plants.first()
        
        if user.assigned_zones.count() == 1 and not form.cleaned_data.get('zone'):
            incident.zone = user.assigned_zones.first()

        if user.assigned_locations.count() == 1 and not form.cleaned_data.get('location'):
            incident.location = user.assigned_locations.first()

        if user.assigned_sublocations.count() == 1 and not form.cleaned_data.get('sublocation'):
            incident.sublocation = user.assigned_sublocations.first()

        # Handle JSON fields from hidden inputs
        incident.affected_body_parts = json.loads(self.request.POST.get('affected_body_parts_json', '[]'))
        incident.unsafe_acts = json.loads(self.request.POST.get('unsafe_acts_json', '[]'))
        incident.unsafe_conditions = json.loads(self.request.POST.get('unsafe_conditions_json', '[]'))
        incident.unsafe_acts_other = self.request.POST.get('unsafe_acts_other', '').strip()
        incident.unsafe_conditions_other = self.request.POST.get('unsafe_conditions_other', '').strip()
        
        incident.save()
        self.object = incident
        form.save_m2m()

        # Handle photo uploads
        photos = self.request.FILES.getlist('photos')
        for photo in photos:
            compressed_photo = compress_image(photo)
            IncidentPhoto.objects.create(
                incident=incident,
                photo=compressed_photo,
                photo_type='INCIDENT_SCENE',
                uploaded_by=self.request.user
            )
        
        # ===== ADD NOTIFICATION HERE - AFTER INCIDENT IS SAVED =====
        print("\n\n" + "#" * 70)
        print("VIEW: INCIDENT SAVED SUCCESSFULLY")
        print("#" * 70)
        print(f"Report Number: {self.object.report_number}")
        print(f"Incident ID: {self.object.id}")
        print(f"Plant: {self.object.plant}")
        print(f"Location: {self.object.location}")
        print("#" * 70)
        print("\nVIEW: Calling notify_incident_reported()...")

        try:
            # ✅ NEW: Use NotificationService instead of old notification_utils
            from apps.notifications.services import NotificationService
            
            NotificationService.notify(
                content_object=self.object,
                notification_type='INCIDENT_REPORTED',
                module='INCIDENT'
            )
            
            print("\nVIEW: ✅ Notifications sent successfully")
        except Exception as e:
            print(f"\nVIEW: ❌ ERROR in notification system: {e}")
            import traceback
            traceback.print_exc()
    
        print("\n" + "#" * 70 + "\n\n")
        
        messages.success(
            self.request,
            f'Incident {self.object.report_number} reported successfully! Investigation required within 7 days.'
        )
        
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

class IncidentDetailView(LoginRequiredMixin, DetailView):
    """View incident details"""
    model = Incident
    template_name = 'accidents/incident_detail.html'
    context_object_name = 'incident'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['photos'] = self.object.photos.all()
        context['action_items'] = self.object.action_items.all()
        
        try:
            context['investigation_report'] = self.object.investigation_report
        except:
            context['investigation_report'] = None
        
        return context


# ============================================================================
# REPLACE YOUR EXISTING IncidentUpdateView WITH THIS UPDATED VERSION
# ============================================================================

def get_zones_by_plant(request, plant_id):
    """
    Fetch zones for a given plant ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
        
    try:
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'zones': list(zones)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_locations_by_zone(request, zone_id):
    """
    Fetch locations for a given zone ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'locations': list(locations)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_sublocations_by_location(request, location_id):
    """
    Fetch sub-locations for a given location ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
        
    try:
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'sublocations': list(sublocations)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
class IncidentUpdateView(LoginRequiredMixin, UpdateView):
    """Update incident report"""
    model = Incident
    form_class = IncidentUpdateForm
    template_name = 'accidents/incident_update.html'
    
    def get_success_url(self):
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        """
        Passes the current request's user to the form's __init__ method.
        This is CRUCIAL for the form logic to work correctly.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        """
        Adds the user's location assignments and JSON data to the template context.
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Pass QuerySets of assigned locations to the template
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)

        # This logic helps the template decide if there's only one option to show
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(is_active=True, location=location)
                else:
                    context['user_assigned_sublocations'] = user.assigned_sublocations.none()
            else:
                context['user_assigned_locations'] = user.assigned_locations.none()
                context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        else:
            context['user_assigned_zones'] = user.assigned_zones.none()
            context['user_assigned_locations'] = user.assigned_locations.none()
            context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        
        # Add departments for the affected person dropdown
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        
        context['active_incident_types'] = IncidentType.objects.filter(is_active=True)

        # ✅ START: ADDED CODE
        # Add JSON-stringified data for safe JS initialization in the template
        context['affected_body_parts_json'] = json.dumps(
            self.object.affected_body_parts or []
        )
        # ✅ END: ADDED CODE
        
        return context
    
        
    #Incident Update Form
    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        incident = self.object  
        form.initial['affected_body_parts_json'] = json.dumps(
            incident.affected_body_parts or []
        )

        form.initial['unsafe_acts_json'] = json.dumps(
            incident.unsafe_acts or []
        )

        form.initial['unsafe_conditions_json'] = json.dumps(
            incident.unsafe_conditions or []
        )
        form.initial['unsafe_acts_other'] = incident.unsafe_acts_other or ''
        form.initial['unsafe_conditions_other'] = incident.unsafe_conditions_other or ''
        if incident.affected_person:
            form.initial['affected_person_id'] = incident.affected_person.id

        return form

    
    def form_valid(self, form):
        # Handle affected person selection
        incident_type = form.cleaned_data.get('incident_type')
        if incident_type:
            form.instance.incident_type = incident_type

        affected_person_id = self.request.POST.get('affected_person_id', '').strip()
        if affected_person_id:
            try:
                affected_employee = User.objects.select_related('department').get(
                    id=affected_person_id,
                    is_active=True
                )
                form.instance.affected_person = affected_employee
                form.instance.affected_person_name = affected_employee.get_full_name()
                form.instance.affected_person_employee_id = affected_employee.employee_id or ''
                form.instance.affected_person_department = affected_employee.department
            except (User.DoesNotExist, ValueError):
                pass
        
        # Handle affected body parts JSON
        affected_body_parts_json = self.request.POST.get('affected_body_parts_json', '[]')
        try:
            form.instance.affected_body_parts = json.loads(affected_body_parts_json)
        except:
            form.instance.affected_body_parts = []
        
        # Handle unsafe acts JSON
        unsafe_acts_json = self.request.POST.get('unsafe_acts_json', '[]')
        try:
            form.instance.unsafe_acts = json.loads(unsafe_acts_json)
        except:
            form.instance.unsafe_acts = []
        
        # Handle unsafe acts other explanation
        form.instance.unsafe_acts_other = self.request.POST.get('unsafe_acts_other', '').strip()
        
        # Handle unsafe conditions JSON
        unsafe_conditions_json = self.request.POST.get('unsafe_conditions_json', '[]')
        try:
            form.instance.unsafe_conditions = json.loads(unsafe_conditions_json)
        except:
            form.instance.unsafe_conditions = []
        
        # Handle unsafe conditions other explanation
        form.instance.unsafe_conditions_other = self.request.POST.get('unsafe_conditions_other', '').strip()
        
        # Handle photo uploads
        photos = self.request.FILES.getlist('photos')
        for photo in photos:
            compressed_photo = compress_image(photo)
            IncidentPhoto.objects.create(
                incident=self.object,
                photo=compressed_photo,
                photo_type='INCIDENT_SCENE',
                uploaded_by=self.request.user
            )
        
        messages.success(self.request, f'Incident {self.object.report_number} updated successfully!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        print("Form errors:", form.errors)
        return super().form_invalid(form)

import json
from django.forms import inlineformset_factory
from django.contrib.auth import get_user_model


User = get_user_model()
class InvestigationReportCreateView(LoginRequiredMixin, CreateView):
    """Create investigation report and its associated action items"""
    model = IncidentInvestigationReport
    form_class = IncidentInvestigationReportForm
    template_name = 'accidents/investigation_report_create.html'

    def dispatch(self, request, *args, **kwargs):
        """Ensure the incident exists before proceeding."""
        self.incident = get_object_or_404(Incident, pk=self.kwargs['incident_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        """Add the incident and action item formset to the context."""
        context = super().get_context_data(**kwargs)
        context['incident'] = self.incident
        
        ActionItemFormSet = inlineformset_factory(
            Incident, 
            IncidentActionItem, 
            form=IncidentActionItemForm, 
            extra=1,
            can_delete=False,
            exclude=['responsible_person'] # This is correct
        )
        
        if self.request.POST:
            context['action_item_formset'] = ActionItemFormSet(self.request.POST, instance=self.incident, prefix='actionitems')
        else:
            context['action_item_formset'] = ActionItemFormSet(instance=self.incident, prefix='actionitems')
            
        return context

    def form_valid(self, form):
        """
        Process the main form and the action item formset.
        This version is more robust for handling formset processing.
        """
        context = self.get_context_data()
        action_item_formset = context['action_item_formset']

        # First, validate both the main form and the formset
        if not action_item_formset.is_valid():
            messages.error(self.request, "There was an error with the action items. Please check the details.")
            return self.form_invalid(form)

        # 1. Save the main investigation report
        investigation = form.save(commit=False)
        investigation.investigator = self.request.user
        investigation.completed_by = self.request.user
        investigation.incident = self.incident
        
        # Process personal and job factors from hidden inputs
        personal_factors_json = self.request.POST.get('personal_factors_json', '[]')
        job_factors_json = self.request.POST.get('job_factors_json', '[]')
        try:
            investigation.personal_factors = json.loads(personal_factors_json)
            investigation.job_factors = json.loads(job_factors_json)
        except json.JSONDecodeError:
            investigation.personal_factors = []
            investigation.job_factors = []
        
        investigation.save()

        # 2. Process and save each action item form individually
        # We iterate over the forms in the formset directly.
        for action_form in action_item_formset:
            # We only process forms that have data entered into them
            if action_form.has_changed():
                action_item = action_form.save(commit=False)
                action_item.incident = self.incident
                action_item.save()  # Save the action item instance to get an ID

                # Get the emails from this specific form's cleaned data
                emails_string = action_form.cleaned_data.get('responsible_person_emails')
                
                if emails_string:
                    # Split the string of emails into a list
                    email_list = [email.strip() for email in emails_string.split(',') if email.strip()]
                    
                    # Find User objects that match the emails.
                    # Our new form validation guarantees that these users exist.
                    users = User.objects.filter(email__in=email_list)
                    
                    # Set the many-to-many relationship for the responsible persons
                    action_item.responsible_person.set(users)
        
        # Update incident status after investigation is complete
        self.incident.status = 'ACTION_IN_PROGRESS'
        self.incident.investigation_completed_date = investigation.completed_date
        self.incident.save()
        # ===== ADD NOTIFICATION: Investigation Completed =====
        try:
            from apps.notifications.services import NotificationService
            
            # Notify that investigation was completed
            # You'll need to add 'INCIDENT_INVESTIGATION_COMPLETED' to your NotificationMaster
            NotificationService.notify(
                content_object=investigation,
                notification_type='INCIDENT_INVESTIGATION_COMPLETED',
                module='INCIDENT_INVESTIGATION_REPORTED'
            )
            
            print("✅ Investigation completion notifications sent")
        except Exception as e:
            print(f"❌ Error sending investigation notifications: {e}")

        messages.success(self.request, "Investigation report and action items have been created successfully.")
        return redirect(self.get_success_url())

        messages.success(self.request, "Investigation report and action items have been created successfully.")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        """
        If the main form is invalid, or if the formset is invalid (handled in form_valid),
        re-render the page with appropriate error messages.
        """
        # The error message for an invalid formset is now added in form_valid.
        # This will show errors from our new email validation.
        messages.error(self.request, "Please correct the errors in the form below.")
        return super().form_invalid(form)
        
    def get_success_url(self):
        """Return the URL to redirect to after successful submission."""
        return reverse('accidents:incident_detail', kwargs={'pk': self.incident.pk})

    def form_invalid(self, form):
        """
        If the form is invalid, re-render the page with error messages.
        This also handles the case where the action_item_formset might be invalid.
        """
        context = self.get_context_data()
        action_item_formset = context['action_item_formset']
        
        # Add a generic error message if the formset is the one with errors
        if not action_item_formset.is_valid():
             messages.error(self.request, "Please correct the errors in the action items section below.")

        return super().form_invalid(form)
        
    def get_success_url(self):
        """Return the URL to redirect to after successful submission."""
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.incident.pk})
    
class ActionItemCreateView(LoginRequiredMixin, CreateView):
    """Create action item"""
    model = IncidentActionItem
    form_class = IncidentActionItemForm
    template_name = 'accidents/action_item_create.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.incident = get_object_or_404(Incident, pk=self.kwargs['incident_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident'] = self.incident
        return context
    
    def form_valid(self, form):
        form.instance.incident = self.incident
        messages.success(self.request, 'Action item created successfully!')
        action_item = form.save()
         # ===== ADD NOTIFICATION: Action Item Assigned =====
        try:
            from apps.notifications.services import NotificationService
            
            # Notify responsible persons about the action item
            NotificationService.notify(
                content_object=action_item,
                notification_type='INCIDENT_ACTION_ASSIGNED',
                module='INCIDENT_ACTION'
            )
            
            print("✅ Action assignment notifications sent")
        except Exception as e:
            print(f"❌ Error sending action assignment notifications: {e}")
        
        messages.success(self.request, 'Action item created successfully!')
        return super().form_valid(form)
    def get_success_url(self):
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.incident.pk})


# AJAX Views
class GetZonesForPlantAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get zones for selected plant"""
    
    def get(self, request, *args, **kwargs):
        plant_id = request.GET.get('plant_id')
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse(list(zones), safe=False)


class GetLocationsForZoneAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get locations for selected zone"""
    
    def get(self, request, *args, **kwargs):
        zone_id = request.GET.get('zone_id')
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse(list(locations), safe=False)
    
class GetSublocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get sublocations for selected location"""
    
    def get(self, request, *args, **kwargs):
        location_id = request.GET.get('location_id')
        from apps.organizations.models import SubLocation
        sublocations = SubLocation.objects.filter(
            location_id=location_id, 
            is_active=True
        ).values('id', 'name', 'code')
        return JsonResponse(list(sublocations), safe=False)

    
from django.views import View

class IncidentPDFDownloadView(LoginRequiredMixin, View):
    """Generate PDF report for incident"""
    
    def get(self, request, pk):
        incident = get_object_or_404(Incident, pk=pk)
        
        # Check permissions
        if not (request.user.is_superuser or 
                request.user == incident.reported_by or
                request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']):
            messages.error(request, "You don't have permission to view this report")
            return redirect('accidents:incident_list')
        
        return generate_incident_pdf(incident)  
    




class IncidentAccidentDashboardView(LoginRequiredMixin, TemplateView):
    """Incident Management Dashboard with Analytics and Filters"""
    template_name = 'accidents/accidents_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        today = timezone.now().date() # Use timezone aware date
        user = self.request.user

        # ==================================================
        # GET FILTER PARAMETERS FROM REQUEST
        # ==================================================
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')  # Format: YYYY-MM

        # ==================================================
        # BASE QUERYSET - Filter by user role
        # ==================================================
        if user.is_superuser or getattr(user, 'role', None) == 'ADMIN':
            incidents = Incident.objects.all()
        elif getattr(user, 'plant', None):
            incidents = Incident.objects.filter(plant=user.plant)
        else:
            incidents = Incident.objects.filter(reported_by=user)

        # ==================================================
        # APPLY FILTERS TO QUERYSET
        # ==================================================
        if selected_plant:
            incidents = incidents.filter(plant_id=selected_plant)
        if selected_zone:
            incidents = incidents.filter(zone_id=selected_zone)
        if selected_location:
            incidents = incidents.filter(location_id=selected_location)
        if selected_sublocation:
            incidents = incidents.filter(sublocation_id=selected_sublocation)
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                incidents = incidents.filter(
                    incident_date__year=year,
                    incident_date__month=month
                )
            except ValueError:
                pass  # Invalid format, ignore filter

        # ==================================================
        # POPULATE FILTER DROPDOWNS
        # ==================================================
        if user.is_superuser or getattr(user, 'role', None) == 'ADMIN':
            all_plants = Plant.objects.filter(is_active=True).order_by('name')
        elif getattr(user, 'plant', None):
            all_plants = Plant.objects.filter(id=user.plant.id, is_active=True)
        else:
            all_plants = Plant.objects.none()
        
        context['plants'] = all_plants

        if selected_plant:
            context['zones'] = Zone.objects.filter(plant_id=selected_plant, is_active=True).order_by('name')
            if selected_zone:
                context['locations'] = Location.objects.filter(zone_id=selected_zone, is_active=True).order_by('name')
                if selected_location:
                    context['sublocations'] = SubLocation.objects.filter(location_id=selected_location, is_active=True).order_by('name')
                else:
                    context['sublocations'] = SubLocation.objects.filter(location__zone_id=selected_zone, is_active=True).order_by('name')
            else:
                context['locations'] = Location.objects.filter(zone__plant_id=selected_plant, is_active=True).order_by('name')
                context['sublocations'] = SubLocation.objects.filter(location__zone__plant_id=selected_plant, is_active=True).order_by('name')
        else:
            context['zones'] = Zone.objects.filter(plant__in=all_plants, is_active=True).order_by('name')
            context['locations'] = Location.objects.filter(zone__plant__in=all_plants, is_active=True).order_by('name')
            context['sublocations'] = SubLocation.objects.filter(location__zone__plant__in=all_plants, is_active=True).order_by('name')

        month_options = []
        for i in range(12):
            # Correctly calculate previous months
            current_month = today.month - i
            current_year = today.year
            if current_month <= 0:
                current_month += 12
                current_year -= 1
            date = datetime.date(current_year, current_month, 1)
            month_options.append({
                'value': date.strftime('%Y-%m'),
                'label': date.strftime('%B %Y')
            })
        context['month_options'] = month_options

        # ==================================================
        # STORE SELECTED FILTER VALUES AND NAMES
        # ==================================================
        context['selected_plant'] = selected_plant
        context['selected_zone'] = selected_zone
        context['selected_location'] = selected_location
        context['selected_sublocation'] = selected_sublocation
        context['selected_month'] = selected_month

        # Get names for active filter display
        context['selected_plant_name'] = ''
        context['selected_zone_name'] = ''
        context['selected_location_name'] = ''
        context['selected_sublocation_name'] = ''
        context['selected_month_label'] = ''

        if selected_plant:
            try:
                plant = Plant.objects.get(id=selected_plant)
                context['selected_plant_name'] = plant.name
            except:
                pass

        if selected_zone:
            try:
                zone = Zone.objects.get(id=selected_zone)
                context['selected_zone_name'] = zone.name
            except:
                pass

        if selected_location:
            try:
                location = Location.objects.get(id=selected_location)
                context['selected_location_name'] = location.name
            except:
                pass

        if selected_sublocation:
            try:
                sublocation = SubLocation.objects.get(id=selected_sublocation)
                context['selected_sublocation_name'] = sublocation.name
            except:
                pass

        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                date_obj = datetime.date(year, month, 1)
                context['selected_month_label'] = date_obj.strftime('%B %Y')
            except:
                pass

        # Flag if any filters are active
        context['has_active_filters'] = bool(
            selected_plant or selected_zone or selected_location or 
            selected_sublocation or selected_month
        )

        # ==================================================
        # BASIC STATISTICS (with filters applied)
        # ==================================================
        context['total_incidents'] = incidents.count()
        context['open_incidents'] = incidents.exclude(status='CLOSED').count()

        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                context['this_month_incidents'] = incidents.filter(
                    incident_date__year=year,
                    incident_date__month=month
                ).count()
                context['current_month_name'] = datetime.date(year, month, 1).strftime('%B')
                context['current_year'] = year
            except:
                context['this_month_incidents'] = incidents.filter(
                    incident_date__month=today.month,
                    incident_date__year=today.year
                ).count()
                context['current_month_name'] = today.strftime('%B')
                context['current_year'] = today.year
        else:
            context['this_month_incidents'] = incidents.filter(
                incident_date__month=today.month,
                incident_date__year=today.year
            ).count()
            context['current_month_name'] = today.strftime('%B')
            context['current_year'] = today.year
        
        context['investigation_pending'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True
        ).count()

        # ==================================================
        # INCIDENT TYPE COUNTS (For Doughnut Chart)
        # ==================================================
        context['lti_count'] = incidents.filter(incident_type__code='LTI').count()
        context['mtc_count'] = incidents.filter(incident_type__code='MTC').count()
        context['fa_count']  = incidents.filter(incident_type__code='FA').count()
        context['hlfi_count'] = incidents.filter(incident_type__code='HLFI').count()


        # ==================================================
        # RECENT INCIDENTS & OVERDUE INVESTIGATIONS
        # ==================================================
        context['recent_incidents'] = incidents.select_related(
            'plant', 'location', 'reported_by'
        ).order_by('-reported_date')[:10]
        context['overdue_investigations'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True,
            investigation_deadline__lt=today
        )

        # ==================================================
        # MONTHLY TREND (Last 6 months from filtered data)
        # ==================================================
        # ... (monthly trend logic remains the same)
        six_months_ago = today - datetime.timedelta(days=180)

        monthly_incidents = incidents.filter(
            incident_date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('incident_date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        monthly_labels = [item['month'].strftime('%b %Y') for item in monthly_incidents]
        monthly_data = [item['count'] for item in monthly_incidents]
        
        context['monthly_labels'] = json.dumps(monthly_labels)
        context['monthly_data'] = json.dumps(monthly_data)

        # ==================================================
        # ****FIXED SECTION****: SEVERITY DISTRIBUTION
        # The model does not have 'severity_level', so we use 'incident_type' instead.
        # This data is for the bar chart named 'severityChart' in the template.
        # ==================================================
        incident_types = IncidentType.objects.order_by('id')

# Count incidents per type
        type_counts = (
            incidents
            .values('incident_type__id')
            .annotate(count=Count('id'))
        )

        # Map: incident_type_id -> count
        count_map = {
            item['incident_type__id']: item['count']
            for item in type_counts
        }

        # Build chart data
        severity_labels = []
        severity_data = []

        for itype in incident_types:
            severity_labels.append(itype.name)
            severity_data.append(count_map.get(itype.id, 0))

        context['severity_labels'] = json.dumps(severity_labels)
        context['severity_data'] = json.dumps(severity_data)
        # ==================================================
        # STATUS DISTRIBUTION
        # ==================================================
        # ... (status distribution logic remains the same)

        status_distribution = incidents.values('status').annotate(
            count=Count('id')
        ).order_by('-count')

        status_labels = []
        status_data = []
        
        # Create a dictionary to hold status display names
        status_choices_dict = dict(Incident.STATUS_CHOICES)

        for item in status_distribution:
            # Append the display name for the label (e.g., "In Progress")
            status_labels.append(status_choices_dict.get(item['status'], item['status']))
            
            # Create the filter URL for the incident list page
            # This will generate a URL like: /accidents/incidents/?status=IN_PROGRESS
            filter_params = {'status': item['status']}
            list_url = reverse('accidents:incident_list') + '?' + urlencode(filter_params)
            
            # Store both the count and the URL for the JavaScript
            status_data.append({
                'count': item['count'],
                'url': list_url
            })

        context['status_labels'] = json.dumps(status_labels)
        # We now pass the more detailed list of objects to the template
        context['status_data'] = json.dumps(status_data)

        # ==================================================
        # MONTH-OVER-MONTH % CHANGE
        # ==================================================
        # ... (month-over-month logic remains the same)
        last_month_start = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        last_month_count = incidents.filter(
            incident_date__year=last_month_start.year,
            incident_date__month=last_month_start.month
        ).count()

        current_month_count = context.get('this_month_incidents', 0)

        if last_month_count > 0:
            change = (
                (context['this_month_incidents'] - last_month_count)
                / last_month_count
            ) * 100
            change = round(change, 1) 
        else:
            change = 0

        context['total_incidents_change'] = change
        context['total_incidents_change_abs'] = abs(change)


        return context    

######################Closure 

class IncidentClosureCheckView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Pre-closure verification page.
    Also handles the attachment upload directly on this page.
    """
    template_name = 'accidents/incident_closure_check.html'
    
    def test_func(self):
        """Check if user has permission to view this page."""
        return (
            self.request.user.is_superuser or
            # self.request.user.can_close_incidents or # Uncomment if you have this on your user model
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']
        )
    
    def get_context_data(self, **kwargs):
        """Helper method to gather all context data."""
        context = {}
        incident = get_object_or_404(Incident, pk=self.kwargs['pk'])
        
        # --- THIS IS THE CORRECTED LINE ---
        # Removed the parentheses from incident.can_be_closed
        can_close, message = incident.can_be_closed
        # ------------------------------------
        
        try:
            investigation = incident.investigation_report
        except IncidentInvestigationReport.DoesNotExist:
            investigation = None
        
        action_items = incident.action_items.all()
        pending_actions = action_items.exclude(status='COMPLETED')
        completed_actions = action_items.filter(status='COMPLETED')
        
        context.update({
            'incident': incident,
            'can_close': can_close,
            'closure_message': message,
            'investigation': investigation,
            'action_items': action_items,
            'pending_actions': pending_actions,
            'completed_actions': completed_actions,
            'total_actions': action_items.count(),
            'completed_count': completed_actions.count(),
        })
        
        if 'attachment_form' not in kwargs:
             context['attachment_form'] = IncidentAttachmentForm(instance=incident)

        return context

    def get(self, request, *args, **kwargs):
        """Handles the display of the verification page."""
        context = self.get_context_data()
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Handles the file upload form submission."""
        incident = get_object_or_404(Incident, pk=self.kwargs['pk'])
        form = IncidentAttachmentForm(request.POST, request.FILES, instance=incident)

        if form.is_valid():
            form.save()
            messages.success(request, 'Closure attachment has been successfully uploaded.')
            return redirect('accidents:incident_closure_check', pk=incident.pk)
        else:
            messages.error(request, 'There was an error uploading the file. Please try again.')
            context = self.get_context_data(attachment_form=form)
            return render(request, self.template_name, context)
        

class IncidentClosureView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Close an incident"""
    model = Incident
    form_class = IncidentClosureForm
    template_name = 'accidents/incident_closure.html'
    
    def test_func(self):
        """Check if user has permission to close incidents"""
        return (
            self.request.user.is_superuser or
            self.request.user.can_close_incidents or
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incident = self.object
        
        # Verify closure eligibility
        can_close, message = incident.can_be_closed
        
        context.update({
            'incident': incident,
            'can_close': can_close,
            'closure_message': message,
        })
        
        return context
    
    def form_valid(self, form):
        incident = form.instance
        
        # Verify incident can be closed
        can_close, message = incident.can_be_closed
        
        if not can_close:
            messages.error(self.request, f"Cannot close incident: {message}")
            return redirect('accidents:incident_detail', pk=incident.pk)
        
        # Set closure details
        incident.closure_date = timezone.now()
        incident.closed_by = self.request.user
        incident.status = 'CLOSED'
        
        messages.success(
            self.request,
            f'Incident {incident.report_number} has been successfully closed.'
        )
        
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.object.pk})


class IncidentReopenView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Reopen a closed incident"""
    
    def test_func(self):
        return (
            self.request.user.is_superuser or
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER']
        )
    
    def post(self, request, pk):
        incident = get_object_or_404(Incident, pk=pk)
        
        if incident.status != 'CLOSED':
            messages.error(request, "Only closed incidents can be reopened")
            return redirect('accidents:incident_detail', pk=pk)
        
        # Reopen incident
        incident.status = 'UNDER_INVESTIGATION'
        incident.closure_date = None
        incident.closed_by = None
        incident.save()
        
        messages.warning(
            request,
            f'Incident {incident.report_number} has been reopened for further investigation.'
        )
        
        return redirect('accidents:incident_detail', pk=pk)    
    
class InvestigationDetailView(LoginRequiredMixin, DetailView):
    """View investigation report details"""
    model = IncidentInvestigationReport
    template_name = 'accidents/investigation_detail.html'
    context_object_name = 'investigation'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident'] = self.object.incident
        return context    
    


##################notification 
class NotificationListView(LoginRequiredMixin, ListView):
    """List user's notifications"""
    model = IncidentNotification
    template_name = 'accidents/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        return IncidentNotification.objects.filter(
            recipient=self.request.user
        ).select_related('incident', 'incident__plant', 'incident__location')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = self.get_queryset().filter(is_read=False).count()
        return context


class MarkNotificationReadView(LoginRequiredMixin, View):
    """Mark notification as read"""
    
    def post(self, request, pk):
        notification = get_object_or_404(
            IncidentNotification, 
            pk=pk, 
            recipient=request.user
        )
        notification.mark_as_read()
        return JsonResponse({'status': 'success'})
#########################################33

class MarkAllNotificationsReadView(LoginRequiredMixin, View):
    """Mark all notifications as read"""
    
    def post(self, request):
        IncidentNotification.objects.filter(
            recipient=request.user,
            is_read=False
        ).update(is_read=True, read_at=timezone.now())
        return JsonResponse({'status': 'success'})   




class IncidentFilterMixin:
    """
    A mixin to provide a filtered queryset of incidents based on URL parameters.
    This can be reused by the Dashboard, Export views, etc.
    """
    def get_filtered_queryset(self):
        user = self.request.user
        
        # Get filter parameters from the URL
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')
        selected_type = self.request.GET.get('type', '')
        selected_status = self.request.GET.get('status', '')

        # Base queryset based on user's role
        if user.is_superuser or getattr(user, 'role', None) == 'ADMIN':
            base_incidents = Incident.objects.select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            ).all()
        elif getattr(user, 'plant', None):
            base_incidents = Incident.objects.filter(plant=user.plant).select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            )
        else:
            base_incidents = Incident.objects.filter(reported_by=user).select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            )

        # Apply filters
        incidents = base_incidents
        if selected_plant:
            incidents = incidents.filter(plant_id=selected_plant)
        if selected_zone:
            incidents = incidents.filter(zone_id=selected_zone)
        if selected_location:
            incidents = incidents.filter(location_id=selected_location)
        if selected_sublocation:
            incidents = incidents.filter(sublocation_id=selected_sublocation)
        if selected_type:
            incidents = incidents.filter(incident_type=selected_type)
        if selected_status == 'open':
            incidents = incidents.exclude(status='CLOSED')
        elif selected_status:
            incidents = incidents.filter(status=selected_status)

        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                incidents = incidents.filter(incident_date__year=year, incident_date__month=month)
            except (ValueError, TypeError):
                pass
        
        return incidents.order_by('-incident_date', '-incident_time')
    

class ExportIncidentsExcelView(LoginRequiredMixin, IncidentFilterMixin, View):
    """
    Handles the export of filtered incident data to an attractive and well-formatted Excel file.
    """
    def get(self, request, *args, **kwargs):
        queryset = self.get_filtered_queryset()
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Incident Report'

        # --- Define Styles ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        row_fills = [
            PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]

        status_fills = {
            'Open': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'In Progress': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'Closed': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        }

        # New style for wrapping text in specific columns
        wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # --- Headers ---
        headers = [
            'Report Number', 'Incident Type', 'Status', 'Incident Date', 'Incident Time',
            'Plant', 'Zone', 'Location', 'Sub-Location', 'Description', 'Affected Person',
            'Nature of Injury', 'Reported By', 'Reported Date', 'Investigation Deadline',
            'Closure Date', 'Closed By'
        ]
        sheet.append(headers)

        # Style the header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Data Population and Styling ---
        # Get column index for text wrapping before the loop
        desc_col_idx = headers.index('Description') + 1
        injury_col_idx = headers.index('Nature of Injury') + 1
        
        for row_index, incident in enumerate(queryset, start=2):
            row_data = [
                incident.report_number,
                incident.incident_type.name if incident.incident_type else 'N/A',
                incident.get_status_display(),
                incident.incident_date,
                incident.incident_time,
                incident.plant.name if incident.plant else 'N/A',
                incident.zone.name if incident.zone else 'N/A',
                incident.location.name if incident.location else 'N/A',
                incident.sublocation.name if incident.sublocation else 'N/A',
                incident.description,
                incident.affected_person_name,
                incident.nature_of_injury,
                incident.reported_by.get_full_name() if incident.reported_by else 'N/A',
                incident.reported_date.strftime("%Y-%m-%d %H:%M") if incident.reported_date else None,
                incident.investigation_deadline,
                incident.closure_date.strftime("%Y-%m-%d %H:%M") if incident.closure_date else None,
                incident.closed_by.get_full_name() if incident.closed_by else 'N/A'
            ]
            sheet.append(row_data)

            # Apply alternating row color (zebra striping)
            current_fill = row_fills[(row_index - 2) % 2]
            for cell in sheet[row_index]:
                cell.fill = current_fill
            
            # Apply wrap text alignment to specific cells
            sheet.cell(row=row_index, column=desc_col_idx).alignment = wrap_alignment
            sheet.cell(row=row_index, column=injury_col_idx).alignment = wrap_alignment

        # --- Conditional Formatting for Status ---
        status_column_letter = get_column_letter(headers.index('Status') + 1)
        for status, fill in status_fills.items():
            rule = CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill)
            sheet.conditional_formatting.add(f'{status_column_letter}2:{status_column_letter}{sheet.max_row}', rule)

        # --- Adjust Column Widths ---
        column_widths = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    column_widths[cell.column_letter] = max(
                        (column_widths.get(cell.column_letter, 0), len(str(cell.value)))
                    )

        for col_letter, width in column_widths.items():
            header_name = sheet[f'{col_letter}1'].value
            # Set a fixed width for columns that need text wrapping
            if header_name in ['Description', 'Nature of Injury']:
                sheet.column_dimensions[col_letter].width = 50
            else:
                # Auto-size other columns with a little padding
                sheet.column_dimensions[col_letter].width = width + 2

        # --- HTTP Response ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"Incident_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)

        return response
    

class IncidentCloseView(LoginRequiredMixin, UpdateView):
    """Close an incident"""
    model = Incident
    template_name = 'accidents/incident_close.html'
    fields = ['closure_remarks', 'lessons_learned', 'preventive_measures']
    def form_valid(self, form):
        print(">>>> form_valid called")
        incident = form.save(commit=False)

        # Set closure fields
        incident.status = 'CLOSED'
        incident.closure_date = timezone.now()
        incident.closed_by = self.request.user
        incident.save()
        print(f">>>> Incident {incident.report_number} closed, calling NotificationService")
        # ===== ADD NOTIFICATION: Incident Closed =====
        try:
            from apps.notifications.services import NotificationService
            NotificationService.notify(
                content_object=incident,
                notification_type='INCIDENT_CLOSED',
                module='INCIDENT_CLOSED'
            )
            print("✅ Incident closure notifications sent")

        except Exception as e:
            print(f"❌ Error sending closure notifications: {e}")

        messages.success(self.request,f'Incident {incident.report_number} has been closed successfully.')
        return redirect('accidents:incident_detail', pk=incident.pk)